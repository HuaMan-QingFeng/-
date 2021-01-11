'''
功能：将<gyzdata1.xlsx>表格整理后输出为<gyzdata1_new.xlsx>表格

'''

import xlrd2

# 打开wokbook和定义workshoot

data1 = xlrd2.open_workbook("./gyzdata1.xlsx")
sheet = data1.sheet_by_name("Sheet1")

#存放学生数据表
turnValue = ['ID','Name','City','Gender','Height','C1','C2','C3','C4','C5','C6','C7','C8','C9','C10','Constitution']
stuValue = [turnValue]

# 创建一个for循环迭代读取xlsx文件每行数据的, 从第二行开始是要跳过标题
for r in range(1, sheet.nrows):
    ID = int(sheet.cell(r, 0).value + 202000)
    Name = sheet.cell(r, 1).value
    City = sheet.cell(r, 2).value


# Gender获取学生性别，同时整理性别的不一致性，统一用“boy”或“girl”表示
    Gender = sheet.cell(r, 3).value
    if Gender =='male':
        Gender = 'boy'
    elif Gender == 'female':
        Gender = 'girl'


# Height获取学生身高，整理身高单位的不一致性，统一用“cm”表示
    Height = sheet.cell(r, 4).value
    if Height != '' and float(Height) <= 100:
        Height = str(int(float(Height)*100))


#获取学生成绩
    C1 = sheet.cell(r, 5).value
    C2 = sheet.cell(r, 6).value
    C3 = sheet.cell(r, 7).value
    C4 = sheet.cell(r, 8).value
    C5 = sheet.cell(r, 9).value
    C6 = sheet.cell(r, 10).value
    C7 = sheet.cell(r, 11).value
    C8 = sheet.cell(r, 12).value
    C9 = sheet.cell(r,13).value
    C10 = sheet.cell(r,14).value
    Constitution = sheet.cell(r,15).value

#获取全部数据为alldata
    alldata = [ID,Name,City,Gender,Height,C1,C2,C3,C4,C5,C6,C7,C8,C9,C10,Constitution]

#将学生数据存进stuValue列表
    stuValue.append(alldata)


#excel操作
import openpyxl

class RunExcel():

    # 定义一个写excel表的方法
    def write(self, excel_path, sheetname, value):
        index = len(value)
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sheetname
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
        wb.save(excel_path)
        print("gyzdata1_new.xlsx格式表格写入数据成功！")

        wb.close()

runexcel = RunExcel() #定义一个excel数据对象


#将stuValue数据写入excel表
runexcel.write("./gyzdata1_new.xlsx", "Sheet", stuValue)







"""
功能：将文本<gyzdata2.txt>文件转换为Excel表格的<gyzdata2_new.xlsx>文件
"""

#创建一个excel
yz = openpyxl.Workbook()

#获取当前excel的sheet
sheet = yz.active

#定义一个空值
strdatas = None

#读取上面txt的数据
with open("./gyzdata2.txt","r") as gg :
    strdatas = gg.readlines()#readlines函数返回一个字符串列表，每个元素为gg的一行内容

ID_turn = ''
#行数，初始值为0
i = 0

#遍历strdatas的长度,100行
for x in range(len(strdatas)):
    #对每一行的strdatas数据按“,”切分，获取到每一个学生的信息
    ID,Name,City,Gender,Height,C1,C2,C3,C4,C5,C6,C7,C8,C9,C10,Constitution = strdatas[x].split(",")
    if ID_turn == ID:
        continue
    else:
        ID_turn = ID

    #按行输入学生信息
    row = str(i + 1)
    sheet["A" + row] = ID
    sheet["B" + row] = Name
    sheet["C" + row] = City

    # 将性别更改为统一格式
    if Gender == 'male':
        sheet["D" + row] = 'boy'
    elif Gender == 'female':
        sheet["D" + row] = 'girl'
    else:
        sheet["D" + row] = Gender

    #将身高更改为统一格式
    if Height == 'Height':
        sheet["E" + row] = Height
    elif float(Height) >= 100:
        sheet["E" + row] = float(Height)
    elif float(Height) <= 10:
        sheet["E" + row] = float(Height) * 100

    #按行输入学生成绩
    sheet["F" + row] = C1
    sheet["G" + row] = C2
    sheet["H" + row] = C3
    sheet["I" + row] = C4
    sheet["J" + row] = C5
    sheet["K" + row] = C6
    sheet["L" + row] = C7
    sheet["M" + row] = C8
    sheet["N" + row] = C9
    sheet["O" + row] = C10
    sheet["P" + row] = Constitution.strip()#使用strip（）去掉末尾的换行符\n

    #每录完一个学生，行数加一
    i += 1

#保存excel，后缀要注意
yz.save("./gyzdata2_new.xlsx")
yz.close()

print("文本<gyzdata2.txt>文件转换为Excel表格的<gyzdata2_new.xlsx>文件转换成功！！")



"""
功能：
将之前生成的<gyzdata1_new.xlsx>与<gyzdata2_new.xlsx>合并为一张表<AllData.xlsx>
"""

import pandas as gyz
import xlrd2
import os

#创建一个input对象，参数为写入路径
input = gyz.ExcelWriter('./AllData.xlsx')

#在当前文件目录下寻找合并文件源
os.chdir('./')

#获取两个数据源表的文件路径
path1 = './gyzdata1_new.xlsx'
path2 = './gyzdata2_new.xlsx'

#将数据源1读出为m1
workbook1 = xlrd2.open_workbook(path1)
m1 = gyz.read_excel(path1, sheet_name='Sheet')
#将数据源2读出为m2
workbook2 = xlrd2.open_workbook(path2)
m2 = gyz.read_excel(path2, sheet_name='Sheet')

#合并m1和m2
turn = (m1.combine_first(m2))#使用m2表的数据补全m1空缺内容并临时存放在turn当中
M = gyz.merge(turn,m2,"outer")#将turn与m2进行外连接，将m1原本缺少的学生添加进来
#去重， keep="first" ： 保留重复的第一个值， sort_values参数：1）以‘ID’做排序  2）True为升序
M.drop_duplicates(subset=['ID'], keep='first').sort_values(['ID'],ascending=True).to_excel(input,'Sheet',index=False)

#保存合并表
input.save()
print("合并数据表成功！！")
